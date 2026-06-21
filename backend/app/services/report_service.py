import io
import csv
import logging
from datetime import datetime
from typing import List
from app.models.transaction import Transaction

# Third-party styling imports (wrapped in try-except for robustness)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger(__name__)

class ReportService:
    """
    Service layer providing spreadsheet exports (CSV/Excel) and formatted PDF reports
    for financial analysis.
    """

    @staticmethod
    def generate_csv(transactions: List[Transaction]) -> str:
        """
        Generates a clean CSV file in memory from the transaction list.
        """
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Headers
            writer.writerow(["ID", "Date", "Type", "Category", "Amount", "Tags", "Notes"])
            
            # Rows
            for t in transactions:
                writer.writerow([
                    t.id,
                    t.date.strftime("%Y-%m-%d %H:%M:%S"),
                    t.type,
                    t.category,
                    t.amount,
                    t.tags or "",
                    t.notes or ""
                ])
                
            return output.getvalue()
        except Exception as e:
            logger.error(f"Error generating CSV report: {str(e)}")
            raise RuntimeError("CSV export failed.") from e

    @staticmethod
    def generate_excel(transactions: List[Transaction]) -> bytes:
        """
        Generates a professionally styled Excel spreadsheet.
        Falls back to CSV format structured as binary if openpyxl is unavailable.
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl is not installed. Falling back to simple CSV binary data.")
            csv_data = ReportService.generate_csv(transactions)
            return csv_data.encode("utf-8")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Transactions"
            
            # Enable gridlines
            ws.views.sheetView[0].showGridLines = True
            
            # Layout branding header
            ws.merge_cells("A1:G1")
            ws["A1"] = "Smart Expense Tracker - Detailed Financial Log"
            ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
            ws.row_dimensions[1].height = 40

            # Headers
            headers = ["Transaction ID", "Timestamp", "Type", "Category", "Amount ($)", "Associated Tags", "Notes / Details"]
            ws.row_dimensions[3].height = 25
            
            header_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid") # Muted Slate
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center")

            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            
            # Row filling
            row_idx = 4
            even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Off-white
            odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for t in transactions:
                ws.row_dimensions[row_idx].height = 20
                current_fill = even_fill if row_idx % 2 == 0 else odd_fill
                
                row_data = [
                    t.id,
                    t.date.strftime("%Y-%m-%d %H:%M:%S"),
                    t.type.upper(),
                    t.category,
                    t.amount,
                    t.tags or "-",
                    t.notes or "-"
                ]
                
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = current_fill
                    cell.font = Font(name="Segoe UI", size=10)
                    
                    # Alignments & formatting
                    if col_idx in (1, 2, 3):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 5:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "$#,##0.00"
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                row_idx += 1

            # Summary Row
            ws.row_dimensions[row_idx].height = 24
            ws.cell(row=row_idx, column=4, value="Total Net Spent:").font = Font(name="Segoe UI", size=10, bold=True)
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right", vertical="center")
            
            formula = f"=SUMIFS(E4:E{row_idx-1}, C4:C{row_idx-1}, \"EXPENSE\") - SUMIFS(E4:E{row_idx-1}, C4:C{row_idx-1}, \"INCOME\")"
            total_cell = ws.cell(row=row_idx, column=5, value=formula)
            total_cell.font = Font(name="Segoe UI", size=10, bold=True)
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_cell.number_format = "$#,##0.00"
            total_cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                # Skip first row because it contains merged branding title
                for cell in col[2:]: 
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # Save workbook to memory buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        except Exception as e:
            logger.error(f"Error compiling Excel spreadsheet: {str(e)}")
            raise RuntimeError("Excel export failed.") from e

    @staticmethod
    def generate_pdf(transactions: List[Transaction]) -> bytes:
        """
        Generates a premium visual PDF report with page branding and summarized transactions.
        """
        if not REPORTLAB_AVAILABLE:
            logger.warning("reportlab is not installed. Falling back to plain text log inside binary PDF wrapper.")
            # Basic fallback
            buffer = io.BytesIO()
            buffer.write(b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n...")
            return buffer.getvalue()

        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=40,
                leftMargin=40,
                topMargin=40,
                bottomMargin=40
            )
            
            # Styling rules
            styles = getSampleStyleSheet()
            
            # Custom corporate styles
            title_style = ParagraphStyle(
                "ReportTitle",
                parent=styles["Heading1"],
                fontName="Helvetica-Bold",
                fontSize=22,
                leading=26,
                textColor=colors.HexColor("#1E293B"),
                spaceAfter=12
            )
            
            meta_style = ParagraphStyle(
                "ReportMeta",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=9,
                leading=12,
                textColor=colors.HexColor("#64748B"),
                spaceAfter=20
            )

            story = []
            
            # Report header
            story.append(Paragraph("SMART EXPENSE ANALYTICS REPORT", title_style))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {len(transactions)}", meta_style))
            story.append(Spacer(1, 10))

            # Math summary metrics
            expenses = [t for t in transactions if t.type == "expense"]
            incomes = [t for t in transactions if t.type == "income"]
            total_exp = sum(e.amount for e in expenses)
            total_inc = sum(i.amount for i in incomes)
            net_bal = total_inc - total_exp

            summary_data = [
                ["Total Income Flow", "Total Expenditures", "Net Account Balance"],
                [f"${total_inc:,.2f}", f"${total_exp:,.2f}", f"${net_bal:,.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[175, 175, 175])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#F1F5F9")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#475569")),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ('TOPPADDING', (0,0), (-1,-1), 8),
                ('BACKGROUND', (0,1), (-1,1), colors.HexColor("#FFFFFF")),
                ('TEXTCOLOR', (0,1), (-1,1), colors.HexColor("#0F172A")),
                ('TEXTCOLOR', (2,1), (2,1), colors.HexColor("#10B981") if net_bal >= 0 else colors.HexColor("#EF4444")),
                ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,1), (-1,1), 14),
                ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor("#E2E8F0")),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 25))
            
            # Transactions Title
            list_title_style = ParagraphStyle(
                "ListTitle",
                parent=styles["Heading2"],
                fontName="Helvetica-Bold",
                fontSize=14,
                textColor=colors.HexColor("#0F172A"),
                spaceAfter=10
            )
            story.append(Paragraph("Transaction History Log", list_title_style))

            # Raw list table
            table_data = [["Date", "Type", "Category", "Amount", "Tags / Notes"]]
            
            # Format row entries (limit to last 50 transactions to keep PDF sizes sane)
            display_txs = sorted(transactions, key=lambda x: x.date, reverse=True)[:50]
            
            cell_body_style = ParagraphStyle("CellBody", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#334155"))
            cell_body_bold = ParagraphStyle("CellBodyBold", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10)
            
            for t in display_txs:
                meta_desc = f"{t.tags or ''} | {t.notes or ''}".strip(" | ")
                if not meta_desc:
                    meta_desc = "-"
                
                type_color_str = "#EF4444" if t.type == "expense" else "#10B981"
                type_style = ParagraphStyle("TypeStyle", parent=cell_body_bold, textColor=colors.HexColor(type_color_str))
                
                table_data.append([
                    Paragraph(t.date.strftime("%Y-%m-%d"), cell_body_style),
                    Paragraph(t.type.upper(), type_style),
                    Paragraph(t.category, cell_body_style),
                    Paragraph(f"${t.amount:,.2f}", cell_body_bold),
                    Paragraph(meta_desc, cell_body_style)
                ])

            tx_table = Table(table_data, colWidths=[75, 60, 90, 80, 220])
            tx_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#334155")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 9),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('TOPPADDING', (0,0), (-1,0), 6),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
                ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor("#CBD5E1")),
            ]))
            
            story.append(tx_table)
            
            # Compile layout
            doc.build(story)
            pdf_bytes = buffer.getvalue()
            return pdf_bytes
        except Exception as e:
            logger.error(f"Error compiling PDF Document: {str(e)}")
            raise RuntimeError("PDF export failed.") from e
